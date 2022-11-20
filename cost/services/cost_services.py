from datetime import datetime

from openpyxl import Workbook

from django.db.models import QuerySet
from django.conf import settings

from rest_framework import status
from rest_framework.serializers import ModelSerializer

from cost import models
from cauth.models import CustomUser
from cost.utils import date_utils
from cost.utils.date_utils import get_month_start, get_month_end
from cost.utils.url_utils import get_next_month_url, get_prev_month_url



class ExcelServices:
    month_start: datetime
    month_end: datetime

    costs: QuerySet(models.Cost)

    serizlizer: ModelSerializer

    def __init__(self, queryset: QuerySet(models.Cost), serializer_class: ModelSerializer, month: int = None, year: int = None) -> None:
        self.month_start = date_utils.get_month_start(month=month, year=year)
        self.month_end = date_utils.get_month_end(self.month_start)

        self.costs = queryset.filter(created_at__gte=self.month_start.date(), created_at__lte=self.month_end.date())

        self.serizlizer = serializer_class(self.costs, many=True)

    @property
    def workbook(self) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Costs'

        columns = ['Date', 'Name', 'Amount', 'Category Name', ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for costs in self.erializer.data:
            row_num += 1

            row = [costs['date'], costs['name'], costs['amount'], costs['category_name'], ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        return workbook


class CostServices:
    @staticmethod
    def get_costs(queryset: QuerySet(models.Cost), serializer_class: ModelSerializer, category_id: int = None, month: int=None, year: int=None) -> dict:
        month_start = get_month_start(month=month, year=year)
        month_end = get_month_end(month_start)
        costs = queryset.filter(created_at__gte=month_start, created_at__lte=month_end)

        if category_id:
            costs = costs.filter(category_id=category_id)

        serializer = serializer_class(costs, many=True)

        return serializer.data
