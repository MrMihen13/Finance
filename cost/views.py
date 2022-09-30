import logging

from openpyxl import Workbook

from django.db.models import Sum
from django_filters import rest_framework
from django.conf import settings
from django.http import HttpResponse

from rest_framework import generics, status, response
from rest_framework.permissions import IsAuthenticated

from cost import models, serializers
from cost import permissions as custom_permissions
from cost.utils.date_utils import get_month_end, get_month_start
from cost.utils.url_utils import get_next_month_url, get_prev_month_url

logger = logging.getLogger(__name__)


class CostListApiView(generics.ListAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = serializers.CostListSerializer
    filter_backends = [rest_framework.DjangoFilterBackend, ]
    filterset_fields = ['category_id']

    def get_queryset(self):
        user = self.request.user
        return models.Cost.objects.filter(user_id=user.id)

    def get(self, request, month=None, year=None, *args, **kwargs):
        month_start = get_month_start(month=month, year=year)
        month_end = get_month_end(month_start)
        costs = self.get_queryset().filter(created_at__gte=month_start, created_at__lte=month_end)

        if request.GET.get('category_id'):
            costs = costs.filter(category_id=request.GET.get('category_id'))

        serializer = self.serializer_class(costs, many=True)
        page = self.paginate_queryset(serializer.data)
        data = self.get_paginated_response(page)

        data.data['results'].append(dict(month_name=month_start.strftime('%B')))
        data.data['next_month'] = get_next_month_url(request=request, month_end=month_end)
        data.data['prev_month'] = get_prev_month_url(request=request, month_start=month_start)
        return data


class CostRetrieveUpdateDestroyApiView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated, custom_permissions.IsOwner)
    serializer_class = serializers.CostSerializer

    def get_queryset(self):
        return models.Cost.objects.all()


class CategoryRetrieveUpdateDestroyApiView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated, custom_permissions.IsOwner)
    serializer_class = serializers.CategorySerializer

    def get_queryset(self):
        return models.Category.objects.all()


class CostCreateApiView(generics.CreateAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = serializers.CostSerializer
    queryset = models.Cost.objects.all()

    def post(self, request, *args, **kwargs):
        data = request.data
        serializer = self.serializer_class(data=data)
        if serializer.is_valid():
            serializer.save(user_id=request.user)
        return response.Response(status=status.HTTP_201_CREATED, data=serializer.data)


class CategoryListApiView(generics.ListAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = serializers.CategorySerializer

    def get_queryset(self):
        return models.Category.objects.filter(user_id=self.request.user.id)

    def list(self, request, *args, **kwargs):
        serializer = self.serializer_class(self.get_queryset(), many=True)
        page = self.paginate_queryset(serializer.data)
        return self.get_paginated_response(page)


class CategoryCreateApiView(generics.CreateAPIView):  # TODO Тесты
    permission_classes = (IsAuthenticated,)
    serializer_class = serializers.CategorySerializer
    queryset = models.Category.objects.all()

    def post(self, request, *args, **kwargs):
        if settings.USER_CATEGORY_LIMIT >= self.queryset.filter(user_id=request.user.id).count():
            data = request.data
            serializer = self.serializer_class(data=data)

            if serializer.is_valid():
                serializer.save(user_id=request.user)

            return response.Response(status=status.HTTP_201_CREATED, data=serializer.data)

        return response.Response(status=status.HTTP_423_LOCKED, data={
            'error': 'Category limit exceeded',
            'message': 'Вы создали максимальное количество категорий'
        })


class GetAnalyticsApiView(generics.GenericAPIView):  # TODO Тесты
    """Getting cost`s analytics for month"""
    permission_classes = (IsAuthenticated,)
    serializer_class = serializers.CostSerializer

    def get_queryset(self):
        return models.Cost.objects.filter(user_id=self.request.user.id)

    def get(self, request, month=None, year=None, *args, **kwargs):
        month_start = get_month_start(month=month, year=year)
        month_end = get_month_end(month_start)

        costs = self.get_queryset().filter(created_at__gte=month_start.date(), created_at__lte=month_end)
        categories = models.Category.objects.filter(user_id=self.request.user.id)
        full_amount = costs.aggregate(Sum('amount'))

        data = {
            'links': dict(
                next_month=get_next_month_url(request=request, month_end=month_end),
                prev_month=get_prev_month_url(request=request, month_start=month_start)
            ),
            'results': dict(
                full_amount=full_amount['amount__sum'], month_name=month_start.strftime('%B'), categories=[]
            ),
        }

        if full_amount['amount__sum']:

            for obj in categories:
                category_amount = costs.filter(category_id=obj.id).aggregate(Sum('amount'))

                if category_amount['amount__sum']:
                    percent = round((category_amount['amount__sum'] * 100) / full_amount['amount__sum'])

                    data['results']['categories'].append({
                        'id': obj.id, 'name': obj.name, 'total': category_amount['amount__sum'], 'percent': percent
                    })

        return response.Response(data=data, status=status.HTTP_200_OK)


class UpgradeRatePlanApiView(generics.GenericAPIView):  # TODO Тесты
    ...  # TODO Апгрейд тарифного плана


class ExelExportApiView(generics.GenericAPIView):  # TODO Тесты
    permission_classes = (IsAuthenticated,)
    serializer_class = serializers.ExcelExportSerializer

    def get_queryset(self):
        return models.Cost.objects.filter(user_id=self.request.user.id)

    def get(self, request, month=None, year=None, *args, **kwargs):
        month_start = get_month_start(month=month, year=year)
        month_end = get_month_end(month_start)

        costs = self.get_queryset().filter(created_at__gte=month_start.date(), created_at__lte=month_end)
        serializer = self.serializer_class(costs, many=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Costs'

        columns = ['Date', 'Name', 'Amount', 'Category Name', ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for costs in serializer.data:
            row_num += 1

            row = [costs['date'], costs['name'], costs['amount'], costs['category_name'], ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        _response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        _response['Content-Disposition'] = f'attachment; filename="costs-{month_start.strftime("%B-%Y").lower()}.xlsx"'

        workbook.save(_response)

        return _response
